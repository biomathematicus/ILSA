"""ILSA Running Budget - reproducible builder for ``ILSA_running budget.xlsx``.

Reads the raw bank (Texas Community Bank) and PayPal exports from
``ILSA_full_ledger.xlsx`` and produces a formatted, multi-sheet planning
workbook that mirrors the layout of the sample "Budget Dashboard" figures.

Unlike a static mock-up, every grounded number is derived from the ledger:

* Starting funds          = consolidated reserves through the month *before* the
                            fiscal year start (internal PayPal<->bank sweeps netted
                            out, exactly as ``treasurer.py`` does).
* Income Tracker rows      = the real FY transactions (correct dates and amounts).
                            PayPal redacts donor names, so rows are labelled
                            generically by source/type (no identities invented).
* Actual kids & invoices   = the invoiced (accrual) program model, mapped to the
                            invoice month -- identical to the fiscal report. Kids =
                            group + welcome (LETC) + graduation (GRAD) books mailed;
                            invoice = the full Dollywood bill (books + mailing).

Everything ties to the fiscal report: the same ``ilsa_ledger`` model feeds both,
so the workbook's actuals, per-child cost (~$2.38), and reserve scenarios cannot
disagree with the report. Forward-looking planning inputs that are *not* in the
data (the budgeted enrollment ramp, the $100k fundraising goal) are seeded as
editable cells, consistent with the sample workbook.

Sheets produced (matching the sample tab order):
    Dashboard | Budget Tracker | Income Tracker | Reserve Scenarios |
    Fundraising Planner | Instructions

Usage:
    python running_budget.py                       # FY Aug 2025 - Jul 2026
    python running_budget.py --fy-start 2026-08    # next fiscal year
    python running_budget.py --out "My Budget.xlsx"
"""

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import ilsa_ledger as L
from ilsa_ledger import (
    CAT_GRANT,
    CAT_INTERNAL,
    CAT_RECURRING,
    CAT_REVERSAL,
    SCENARIO_BAND,
    build_monthly,
    horizon_to_end,
    invoice_monthly,
    load_ledger,
    project_period,
    reserve_scenarios,
)

# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #
DATA_FILE = L.DATA_FILE
OUT_FILE = L.ROOT / "ILSA_running_budget.xlsx"

FUNDRAISING_GOAL = 100_000          # planning input (editable on the sheet)

# Seeded budgeted enrollment ramp + DPIL invoice budget (planning layer; matches
# the sample figures). Index 0..11 == Aug..Jul of the fiscal year.
BUDGET_KIDS = [370, 425, 490, 560, 650, 750, 850, 980, 1120, 1300, 1500, 1700]
BUDGET_INVOICE = [900, 1000, 1200, 1350, 1580, 1800, 2100, 2400, 2750, 3200, 3650, 4200]

# Friendly, identity-free labels for PayPal transaction types.
PP_TYPE_LABEL = {
    "Subscription Payment": ("PayPal Subscription", "Recurring"),
    "Donation Payment": ("PayPal Donation", "Individual"),
    "Express Checkout Payment": ("PayPal Donation", "Individual"),
    "Mass Pay Payment": ("PayPal Giving Fund", "Platform"),
    "Website Payment": ("PayPal Website", "Individual"),
}

# --------------------------------------------------------------------------- #
# Palette + styling helpers
# --------------------------------------------------------------------------- #
NAVY = "0E2B3E"
TEAL = "17A2B8"
GREEN = "2E8B57"
GREEN_KPI = "27AE60"
RED = "C0392B"
BLUE = "2E86C1"
GOLD = "F0AD4E"
GOLD_BAR = "FFC107"
YELLOW = "FFF200"
STRIPE = "EAF1F7"
WHITE = "FFFFFF"
ENTRY = "0070C0"          # blue "enter data here" font
GRID = "BFCAD4"

CUR = '$#,##0;($#,##0);"—"'
CUR2 = '$#,##0.00;($#,##0.00);"—"'
NUM = '#,##0;;"—"'
PCT = '0.0%;;"—"'

_THIN = Side(style="thin", color=GRID)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_FY_MONTHS = ["Aug", "Sep", "Oct", "Nov", "Dec", "Jan",
              "Feb", "Mar", "Apr", "May", "Jun", "Jul"]


def _apply(c, fill, color, bold, size, align, valign, wrap, fmt, border, italic):
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.font = Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = _BORDER


def cell(ws, ref, value=None, fill=None, color="000000", bold=False, size=11,
         align="left", valign="center", wrap=False, fmt=None, border=True,
         italic=False):
    c = ws[ref]
    if value is not None:
        c.value = value
    _apply(c, fill, color, bold, size, align, valign, wrap, fmt, border, italic)
    return c


def merged(ws, rng, value=None, **kw):
    """Set value on the top-left cell, style the whole range, then merge it."""
    top = rng.split(":")[0]
    if value is not None:
        ws[top].value = value
    for row in ws[rng]:
        for c in row:
            _apply(c, kw.get("fill"), kw.get("color", "000000"), kw.get("bold", False),
                   kw.get("size", 11), kw.get("align", "left"), kw.get("valign", "center"),
                   kw.get("wrap", False), kw.get("fmt"), kw.get("border", True),
                   kw.get("italic", False))
    ws.merge_cells(rng)


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# --------------------------------------------------------------------------- #
# Ledger consolidation (mirrors treasurer.py so numbers reconcile)
# --------------------------------------------------------------------------- #
def starting_reserves(monthly, fy_start):
    """Consolidated reserves accumulated through the month before ``fy_start``.

    Uses the same inception-to-date cash position as the fiscal report, so the two
    deliverables always agree.
    """
    start = pd.Period(fy_start, "M")
    prior = monthly.loc[monthly.index < start]
    return float(prior["cash_position"].iloc[-1]) if len(prior) else 0.0


def invoice_actuals_by_month(invoice_m, fy_start):
    """Per FY-month index 0..11: invoiced (accrual) program cost and children
    served, from the shared invoice model -- identical to the fiscal report.

    Cost is the invoiced amount for that invoice month (not the bank-payment date),
    and kids are the full book count (group + LETC + GRAD = pieces mailed)."""
    start = pd.Period(fy_start, "M")
    cost, kids = {}, {}
    if invoice_m is None or invoice_m.empty:
        return cost, kids
    for p in invoice_m.index:
        idx = (p - start).n
        if not (0 <= idx < 12):
            continue
        ac = invoice_m.loc[p, "accrual_cost"]
        if pd.notna(ac) and ac > 0:
            cost[idx] = float(ac)
        k = int(invoice_m.loc[p, "kids"])
        if k > 0:
            kids[idx] = k
    return cost, kids


def latest_enrollment(invoice_m):
    """Children served in the most recent invoice month (group + LETC + GRAD)."""
    if invoice_m is None or invoice_m.empty:
        return None
    return int(invoice_m["kids"].iloc[-1])


def income_rows(ledger):
    """Build identity-free income rows for ALL years (the full running ledger).

    Every donation, grant, and transfer from inception to date is listed, sorted
    by date. PayPal giving uses gross (pre-fee) amounts -- the gift the donor made
    -- while the internal PayPal->bank sweep is shown as an informational
    ``Transfer`` row that the Income Tracker's ``Total Raised`` deliberately
    excludes.
    """
    rows = []

    for _, r in ledger.iterrows():
        cat, amt, gross = r["category"], float(r["amount"]), float(r["gross"])
        if r["source"] == "PayPal":
            if cat == CAT_INTERNAL:
                continue  # the bank-side sweep is shown via the TCB transfer row
            label, kind = PP_TYPE_LABEL.get(r["description"], (f"PayPal {r['description']}", "Individual"))
            rows.append({"date": r["date"], "source": label, "type": kind,
                         "amount": gross, "platform": "PayPal", "note": ""})
        elif cat == CAT_RECURRING:          # Bonterra
            rows.append({"date": r["date"], "source": "Bonterra", "type": "Online Platform",
                         "amount": amt, "platform": "Bonterra", "note": ""})
        elif cat == CAT_GRANT:
            note = "Major grant" if amt >= 50_000 else ""
            rows.append({"date": r["date"], "source": "TCB Deposit", "type": "Grant/Major",
                         "amount": amt, "platform": "Bank", "note": note})
        elif cat == CAT_REVERSAL:
            rows.append({"date": r["date"], "source": "TCB Chargeback", "type": "Chargeback",
                         "amount": amt, "platform": "Bank", "note": "Donor-reversed gift"})
        elif cat == CAT_INTERNAL and amt > 0:
            rows.append({"date": r["date"], "source": "TCB <- PayPal sweep", "type": "Transfer",
                         "amount": amt, "platform": "Bank",
                         "note": "Internal transfer - excluded from Total Raised"})

    rows.sort(key=lambda x: x["date"])
    return rows


# --------------------------------------------------------------------------- #
# Sheet builders
# --------------------------------------------------------------------------- #
def build_dashboard(ws, monthly, current_kids):
    """Actual financial position across ALL months (not a single-year budget slice).

    KPIs and the monthly table are real consolidated ledger figures: total income
    and expenses to date, current reserves (cash on hand), and the full month-by-
    month activity from inception.
    """
    set_widths(ws, [16, 16, 16, 16, 16])
    span = (f"{monthly.index.min().strftime('%B %Y')} - "
            f"{monthly.index.max().strftime('%B %Y')}")

    merged(ws, "A1:E2", "IMAGINATION LIBRARY OF SAN ANTONIO - FINANCIAL DASHBOARD",
           fill=NAVY, color=WHITE, bold=True, size=15, align="center")
    merged(ws, "A3:E3", f"All months, inception to date  ·  {span}", fill=TEAL,
           color=WHITE, bold=True, italic=True, size=12, align="center")

    total_income = float(monthly["total_income"].sum())
    total_expense = float(monthly["total_expense"].sum())
    reserves = float(monthly["cash_position"].iloc[-1])

    kpis = [
        ("A5", "A6", "Total Income\n(all years)", round(total_income), GREEN_KPI, CUR),
        ("B5", "B6", "Total Expenses\n(all years)", round(total_expense), RED, CUR),
        ("C5", "C6", "Current Reserves", round(reserves), GREEN_KPI, CUR),
        ("D5", "D6", "Latest Enrollment", current_kids, BLUE, NUM),
        ("E5", "E6", "Months Tracked", len(monthly), NAVY, NUM),
    ]
    for tref, vref, title, value, color, fmt in kpis:
        cell(ws, tref, title, fill=color, color=WHITE, bold=True, size=10,
             align="center", wrap=True)
        cell(ws, vref, value, color=color, bold=True, size=16, align="center", fmt=fmt)
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[6].height = 26

    merged(ws, "A8:E8", "Monthly Activity (all months)", fill=BLUE, color=WHITE,
           bold=True, size=12, align="center")
    headers = ["Month", "Income", "Expenses", "Net", "Reserves"]
    for i, h in enumerate(headers, start=1):
        cell(ws, f"{get_column_letter(i)}9", h, fill=NAVY, color=WHITE, bold=True,
             size=10, align="center")

    first = 10
    for i, (period, row) in enumerate(monthly.iterrows()):
        r = first + i
        stripe = STRIPE if i % 2 else WHITE
        cell(ws, f"A{r}", period.strftime("%b %Y"), fill=stripe, align="center")
        cell(ws, f"B{r}", round(float(row["total_income"]), 2), fill=stripe, align="center", fmt=CUR)
        cell(ws, f"C{r}", round(float(row["total_expense"]), 2), fill=stripe, align="center", fmt=CUR)
        cell(ws, f"D{r}", round(float(row["net"]), 2), fill=stripe, align="center", fmt=CUR)
        cell(ws, f"E{r}", round(float(row["cash_position"]), 2), fill=stripe, align="center", fmt=CUR)

    tr = first + len(monthly)
    cell(ws, f"A{tr}", "TOTAL", fill=NAVY, color=WHITE, bold=True, align="center")
    cell(ws, f"B{tr}", round(total_income, 2), fill=NAVY, color=WHITE, bold=True, align="center", fmt=CUR)
    cell(ws, f"C{tr}", round(total_expense, 2), fill=NAVY, color=WHITE, bold=True, align="center", fmt=CUR)
    cell(ws, f"D{tr}", round(total_income - total_expense, 2), fill=NAVY, color=WHITE, bold=True, align="center", fmt=CUR)
    cell(ws, f"E{tr}", round(reserves, 2), fill=NAVY, color=WHITE, bold=True, align="center", fmt=CUR)


def build_budget_tracker(ws, start_funds, cost_actuals, kids_actuals):
    set_widths(ws, [22, 14, 14, 14, 14, 12, 11])
    merged(ws, "A1:G2", "MONTHLY BUDGET TRACKER  ·  Blue cells = data entry",
           fill=NAVY, color=WHITE, bold=True, size=15, align="center")

    cell(ws, "A3", "Starting Funds (FY open):", bold=True, align="left")
    cell(ws, "B3", start_funds, fill=STRIPE, color=ENTRY, bold=True, align="center", fmt=CUR)
    cell(ws, "A4", "Total Budgeted Costs:", bold=True, align="left")
    cell(ws, "B4", "=SUM(D9:D20)", fill=STRIPE, bold=True, align="center", fmt=CUR)
    cell(ws, "A5", "Fundraising Goal:", bold=True, align="left")
    cell(ws, "B5", FUNDRAISING_GOAL, fill=STRIPE, color=ENTRY, bold=True, align="center", fmt=CUR)
    for ref in ("C3", "C4", "C5"):
        cell(ws, ref, None, border=False)

    merged(ws, "A7:G7", "PROGRAM COSTS — Monthly Enrollment & DPIL Invoices",
           fill=TEAL, color=WHITE, bold=True, size=12, align="center")
    headers = ["Month", "Budget: Kids", "Actual Kids\n(enter)", "Budget: Invoice",
               "Actual Invoice\n(enter)", "Variance $", "% of Budget"]
    for i, h in enumerate(headers, start=1):
        cell(ws, f"{get_column_letter(i)}8", h, fill=NAVY, color=WHITE, bold=True,
             size=10, align="center", wrap=True)
    ws.row_dimensions[8].height = 32

    for i, mon in enumerate(_FY_MONTHS):
        r = 9 + i
        stripe = STRIPE if i % 2 else WHITE
        cost_a = cost_actuals.get(i)
        kids_a = kids_actuals.get(i)
        cell(ws, f"A{r}", mon, fill=NAVY, color=WHITE, bold=True, align="center")
        cell(ws, f"B{r}", BUDGET_KIDS[i], fill=stripe, align="center", fmt=NUM)
        cell(ws, f"C{r}", kids_a, fill=stripe, color=ENTRY, align="center", fmt=NUM)  # invoiced kids
        cell(ws, f"D{r}", BUDGET_INVOICE[i], fill=stripe, align="center", fmt=CUR)
        cell(ws, f"E{r}", round(cost_a, 2) if cost_a is not None else None,
             fill=stripe, color=ENTRY, align="center", fmt=CUR2)
        cell(ws, f"F{r}", f'=IF(E{r}="","",D{r}-E{r})', fill=stripe, align="center", fmt=CUR)
        cell(ws, f"G{r}", f'=IF(OR(E{r}="",D{r}=0),"",E{r}/D{r})', fill=stripe,
             align="center", fmt=PCT)

    r = 21
    cell(ws, f"A{r}", "TOTALS", fill=NAVY, color=WHITE, bold=True, align="center")
    cell(ws, f"B{r}", "=SUM(B9:B20)", fill=NAVY, color=WHITE, bold=True, align="center", fmt=NUM)
    cell(ws, f"C{r}", "=SUM(C9:C20)", fill=NAVY, color=WHITE, bold=True, align="center", fmt=NUM)
    cell(ws, f"D{r}", "=SUM(D9:D20)", fill=NAVY, color=WHITE, bold=True, align="center", fmt=CUR)
    cell(ws, f"E{r}", "=SUM(E9:E20)", fill=NAVY, color=WHITE, bold=True, align="center", fmt=CUR)
    cell(ws, f"F{r}", "=SUM(F9:F20)", fill=NAVY, color=WHITE, bold=True, align="center", fmt=CUR)
    cell(ws, f"G{r}", None, fill=NAVY, border=True)

    merged(ws, "A23:G23", "OPERATIONAL & MARKETING COSTS", fill=GOLD, color=NAVY,
           bold=True, size=12, align="center")
    headers = ["Item", "Budgeted ($)", "Actual ($)\n(enter)", "Frequency", "Notes", "", ""]
    for i, h in enumerate(headers, start=1):
        cell(ws, f"{get_column_letter(i)}24", h, fill=NAVY, color=WHITE, bold=True,
             size=10, align="center", wrap=True)
    ws.row_dimensions[24].height = 30

    ops = [("Mail Box", 350, "yearly", "")]
    for i in range(7):  # one known item + blank entry rows for the Director
        r = 25 + i
        stripe = STRIPE if i % 2 else WHITE
        item = ops[i] if i < len(ops) else ("", None, "", "")
        cell(ws, f"A{r}", item[0] or None, fill=stripe, align="left")
        cell(ws, f"B{r}", item[1], fill=stripe, align="center", fmt=CUR)
        cell(ws, f"C{r}", None, fill=stripe, color=ENTRY, align="center", fmt=CUR)
        cell(ws, f"D{r}", item[2] or None, fill=stripe, align="center", italic=True)
        cell(ws, f"E{r}", item[3] or None, fill=stripe, align="left")
        cell(ws, f"F{r}", None, fill=stripe)
        cell(ws, f"G{r}", None, fill=stripe)


def build_income_tracker(ws, rows):
    set_widths(ws, [11, 24, 14, 13, 11, 8, 26, 11])
    if rows:
        span = (f"{min(r['date'] for r in rows):%b %Y} -- "
                f"{max(r['date'] for r in rows):%b %Y}")
    else:
        span = "no data"
    merged(ws, "A1:H2", f"INCOME & FUNDRAISING TRACKER  ·  All Years ({span})",
           fill=GREEN, color=WHITE, bold=True, size=15, align="center")
    merged(ws, "A3:H3", f"Cumulative Fundraising Goal: ${FUNDRAISING_GOAL:,.0f}",
           fill=GOLD, color=NAVY, bold=True, size=12, align="center")

    last = 9 + len(rows) - 1
    kpis = [
        ("A5:B5", "A6:B6", "Total Raised", f'=SUMIFS(D9:D{last},C9:C{last},"<>Transfer")', GREEN_KPI, CUR),
        ("C5:D5", "C6:D6", "Goal", "='Budget Tracker'!B5", NAVY, CUR),
        ("E5:F5", "E6:F6", "Remaining", f'=C6-A6', RED, CUR),
        ("G5:H5", "G6:H6", "% of Goal", f'=IF(C6=0,"",A6/C6)', BLUE, PCT),
    ]
    for hr, vr, title, formula, color, fmt in kpis:
        merged(ws, hr, title, fill=color, color=WHITE, bold=True, size=11, align="center")
        merged(ws, vr, formula, color=color, bold=True, size=16, align="center", fmt=fmt)
    ws.row_dimensions[6].height = 28

    headers = ["Date", "Donor / Source", "Type", "Amount", "Platform", "Month",
               "Notes", "Receipt #"]
    for i, h in enumerate(headers, start=1):
        cell(ws, f"{get_column_letter(i)}8", h, fill=NAVY, color=WHITE, bold=True,
             size=10, align="center")
    ws.row_dimensions[8].height = 22

    for i, row in enumerate(rows):
        r = 9 + i
        stripe = STRIPE if i % 2 else WHITE
        cell(ws, f"A{r}", row["date"], fill=stripe, align="center", fmt="m/d/yyyy")
        cell(ws, f"B{r}", row["source"], fill=stripe, color=ENTRY, align="left")
        cell(ws, f"C{r}", row["type"], fill=stripe, align="left")
        cell(ws, f"D{r}", row["amount"], fill=stripe, color=ENTRY, align="center", fmt=CUR2)
        cell(ws, f"E{r}", row["platform"], fill=stripe, align="center")
        cell(ws, f"F{r}", row["date"].strftime("%b"), fill=stripe, align="center")
        cell(ws, f"G{r}", row["note"] or None, fill=stripe, align="left")
        cell(ws, f"H{r}", None, fill=stripe, align="center")

    for j in range(4):  # blank entry rows
        r = 9 + len(rows) + j
        stripe = STRIPE if (len(rows) + j) % 2 else WHITE
        for i in range(1, 9):
            cell(ws, f"{get_column_letter(i)}{r}", None, fill=stripe,
                 color=ENTRY if i in (2, 4) else "000000",
                 align="center" if i != 2 else "left",
                 fmt=CUR2 if i == 4 else None)


def build_reserve_scenarios(ws, future, base, pess, opt):
    """Reserve projection scenarios -- identical model to the fiscal report.

    Baseline = linear trend of recurring giving and program cost on past data;
    optimistic/pessimistic vary the baseline net cash flow by +/- SCENARIO_BAND
    per year. Values are computed by the shared ledger model (read-only).
    """
    band = int(round(SCENARIO_BAND * 100))
    set_widths(ws, [16, 20, 18, 20, 4])
    merged(ws, "A1:D2",
           "RESERVE PROJECTION SCENARIOS",
           fill=NAVY, color=WHITE, bold=True, size=14, align="center")
    merged(ws, "A3:D4",
           f"Same model as the fiscal report: baseline = linear trend of recurring giving "
           f"and program cost on past data; optimistic/pessimistic vary the baseline net "
           f"cash flow by +/-{band}% per year. No new grants assumed. Re-run running_budget.py "
           f"to refresh.",
           fill=GOLD, color=NAVY, bold=True, size=10, align="center", wrap=True)
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 18

    hdr = ["Year-end", f"Pessimistic (-{band}%/yr)", "Baseline", f"Optimistic (+{band}%/yr)"]
    for i, h in enumerate(hdr, start=1):
        cell(ws, f"{get_column_letter(i)}6", h, fill=NAVY, color=WHITE, bold=True,
             size=11, align="center")

    last_idx_by_year = {}
    for i, p in enumerate(future):
        last_idx_by_year[p.year] = i

    r = 7
    for j, year in enumerate(sorted(last_idx_by_year)):
        i = last_idx_by_year[year]
        stripe = STRIPE if j % 2 else WHITE
        cell(ws, f"A{r}", future[i].strftime("%b %Y"), fill=stripe, align="center")
        cell(ws, f"B{r}", round(float(pess[i])), fill=stripe, align="center", fmt=CUR)
        cell(ws, f"C{r}", round(float(base[i])), fill=stripe, align="center", fmt=CUR)
        cell(ws, f"D{r}", round(float(opt[i])), fill=stripe, align="center", fmt=CUR)
        r += 1


def build_fundraising_planner(ws, rows):
    set_widths(ws, [26, 20, 13, 13, 12, 14, 12, 30])
    merged(ws, "A1:H2", "FUNDRAISING PIPELINE & GRANT PLANNER", fill=GREEN,
           color=WHITE, bold=True, size=15, align="center")
    merged(ws, "A3:H4",
           "Track every grant, campaign, and donor here. Blue = enter data. Totals update automatically.",
           fill=STRIPE, color=NAVY, italic=True, size=11, align="center")

    headers = ["Source / Grant / Event", "Type", "Target ($)", "Raised ($)",
               "Status", "Expected Date", "Lead", "Notes"]
    for i, h in enumerate(headers, start=1):
        cell(ws, f"{get_column_letter(i)}6", h, fill=NAVY, color=WHITE, bold=True,
             size=10, align="center", wrap=True)
    ws.row_dimensions[6].height = 28

    # Grounded, identity-free pipeline rows derived from the FY ledger totals.
    recurring = sum(r["amount"] for r in rows
                    if r["platform"] == "PayPal" and r["type"] == "Recurring")
    bonterra = sum(r["amount"] for r in rows if r["platform"] == "Bonterra")
    grants = sum(r["amount"] for r in rows if r["type"] == "Grant/Major")
    seed = [
        ("PayPal recurring subscriptions", "Individual Recurring", None, round(recurring, 2),
         "Active", "Monthly", "", "Two monthly subscriptions ($250/mo + $53.56/mo)"),
        ("Bonterra online giving", "Online Platform", None, round(bonterra, 2),
         "Active", "Ongoing", "", "Live platform"),
        ("Bank deposits (grants & major gifts)", "Grant/Major", None, round(grants, 2),
         "Received", "All years to date", "", "Largest single deposit: $110,000"),
    ]
    for i, row in enumerate(seed):
        r = 7 + i
        stripe = STRIPE if i % 2 else WHITE
        cell(ws, f"A{r}", row[0], fill=stripe, align="left")
        cell(ws, f"B{r}", row[1], fill=stripe, color=ENTRY, align="left")
        cell(ws, f"C{r}", row[2], fill=stripe, color=ENTRY, align="center", fmt=CUR)
        cell(ws, f"D{r}", row[3], fill=stripe, color=ENTRY, align="center", fmt=CUR)
        cell(ws, f"E{r}", row[4], fill=stripe, align="center")
        cell(ws, f"F{r}", row[5], fill=stripe, align="center")
        cell(ws, f"G{r}", row[6] or None, fill=stripe, align="center")
        cell(ws, f"H{r}", row[7] or None, fill=stripe, align="left")

    blank_start = 7 + len(seed)
    for j in range(8):  # pipeline rows for the Director to fill
        r = blank_start + j
        stripe = STRIPE if (len(seed) + j) % 2 else WHITE
        for i in range(1, 9):
            cell(ws, f"{get_column_letter(i)}{r}", None, fill=stripe,
                 color=ENTRY if i in (1, 2, 3, 4) else "000000",
                 align="center" if i in (3, 4, 5, 6, 7) else "left",
                 fmt=CUR if i in (3, 4) else None)

    tr = blank_start + 8
    cell(ws, f"A{tr}", "TOTALS", fill=NAVY, color=WHITE, bold=True, align="center")
    cell(ws, f"B{tr}", None, fill=NAVY)
    cell(ws, f"C{tr}", f"=SUM(C7:C{tr-1})", fill=NAVY, color=WHITE, bold=True, align="center", fmt=CUR)
    cell(ws, f"D{tr}", f"=SUM(D7:D{tr-1})", fill=NAVY, color=WHITE, bold=True, align="center", fmt=CUR)
    for col in ("E", "F", "G", "H"):
        cell(ws, f"{col}{tr}", None, fill=NAVY)


def build_instructions(ws):
    set_widths(ws, [16, 52, 18, 22])
    merged(ws, "A1:D2", "HOW TO USE THIS WORKBOOK", fill=NAVY, color=WHITE,
           bold=True, size=15, align="center")

    headers = ["SHEET", "PURPOSE", "WHO UPDATES", "FREQUENCY"]
    for i, h in enumerate(headers, start=1):
        cell(ws, f"{get_column_letter(i)}4", h, fill=NAVY, color=WHITE, bold=True,
             size=10, align="left")
    rows = [
        ("Dashboard", "Auto-calculated summary. KPIs and the monthly table update from the Budget Tracker.",
         "Read only", "Review weekly/monthly"),
        ("Budget Tracker", "Enter actual kids enrolled (col C) and actual DPIL invoice (col E) each month. "
         "Also enter actual operational costs as you spend them.", "Director", "Monthly"),
        ("Income Tracker", "Every donation, grant, or transfer is logged here. Pre-filled from the ledger "
         "(PayPal donor names are redacted in the export, so sources are labelled generically).",
         "Director", "As donations come in"),
        ("Reserve Scenarios", "Projected reserves through 2029 under baseline and +/-15%/yr scenarios. "
         "Same model as the fiscal report; refreshed when you re-run the script.", "Director / Board", "Quarterly planning"),
        ("Fundraising Planner", "Track all fundraising sources: grants, campaigns, recurring donors. "
         "Update status and amounts raised.", "Director", "Monthly or as needed"),
    ]
    for i, row in enumerate(rows):
        r = 5 + i
        stripe = STRIPE if i % 2 else WHITE
        cell(ws, f"A{r}", row[0], fill=stripe, bold=True, align="left", valign="top")
        cell(ws, f"B{r}", row[1], fill=stripe, align="left", wrap=True, valign="top")
        cell(ws, f"C{r}", row[2], fill=stripe, align="left", valign="top")
        cell(ws, f"D{r}", row[3], fill=stripe, align="left", valign="top")
        ws.row_dimensions[r].height = 46

    merged(ws, "A11:D11",
           "COLOR CODING:  Blue text = enter data here   |   Black text = formula (do not edit)   "
           "|   Yellow cell = key assumption you can change",
           fill=GOLD_BAR, color=NAVY, bold=True, size=10, align="center")


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def fy_label_from(fy_start):
    start = pd.Period(fy_start, "M")
    end = start + 11
    return f"{start.strftime('%B %Y')} – {end.strftime('%B %Y')}"


def main():
    p = argparse.ArgumentParser(description="Build the ILSA running-budget workbook from the ledger.")
    p.add_argument("--fy-start", default="2025-08", help="Fiscal-year start month (YYYY-MM). Default 2025-08.")
    p.add_argument("--data", default=str(DATA_FILE), help="Path to ILSA_full_ledger.xlsx.")
    p.add_argument("--out", default=str(OUT_FILE), help="Output workbook path.")
    args = p.parse_args()

    data_path = Path(args.data)
    if not data_path.exists():
        print(f"ERROR: data file not found: {data_path.resolve()}")
        return

    print(f"Loading ledger: {data_path}")
    ledger = load_ledger(data_path)
    monthly = build_monthly(ledger)
    invoice_m = invoice_monthly()
    # Invoiced (accrual) program cost, aligned to the ledger months -- the same
    # series the fiscal report trends, so the reserve scenarios match exactly.
    monthly["program_accrual"] = (
        invoice_m["accrual_cost"].reindex(monthly.index).fillna(0.0)
    )

    start_funds = starting_reserves(monthly, args.fy_start)
    cost_actuals, kids_actuals = invoice_actuals_by_month(invoice_m, args.fy_start)
    rows = income_rows(ledger)  # full running ledger: all years
    fy_label = fy_label_from(args.fy_start)

    proj = project_period(monthly, horizon=horizon_to_end(monthly.index[-1]),
                          expense_col="program_accrual")
    future, base, pess, opt = reserve_scenarios(monthly, proj)

    print(f"  Fiscal year: {fy_label}")
    print(f"  Starting reserves (computed): ${start_funds:,.2f}")
    print(f"  Income ledger rows (all years): {len(rows)}")
    print(f"  Months with invoiced program cost: {sorted(cost_actuals)}")
    print(f"  Baseline reserves at {future[-1]}: ${base[-1]:,.0f} "
          f"(pess ${pess[-1]:,.0f}, opt ${opt[-1]:,.0f})")

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    build_dashboard(ws, monthly, latest_enrollment(invoice_m))
    build_budget_tracker(wb.create_sheet("Budget Tracker"), start_funds,
                         cost_actuals, kids_actuals)
    build_income_tracker(wb.create_sheet("Income Tracker"), rows)
    build_reserve_scenarios(wb.create_sheet("Reserve Scenarios"), future, base, pess, opt)
    build_fundraising_planner(wb.create_sheet("Fundraising Planner"), rows)
    build_instructions(wb.create_sheet("Instructions"))

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False

    out_path = Path(args.out)
    wb.save(out_path)
    print(f"\nWrote {out_path.resolve()}")


if __name__ == "__main__":
    main()
